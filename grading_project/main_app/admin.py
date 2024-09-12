from django.contrib import admin
from django.urls import path
from django.shortcuts import render
from django.utils.html import format_html
from django.urls import reverse
from main_app.models import *
from django.contrib.auth.models import User
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .forms import GradingForm


def user_str(self):
    full_name = self.get_full_name()
    return full_name if full_name else self.username


User.add_to_class("__str__", user_str)


class ProfileAdmin(admin.ModelAdmin):
    list_display = ['full_name', 'get_cathedras', 'ratings', 'get_report_button']
    search_fields = ['user__first_name', 'user__last_name', 'teaching_cathedras__cathedra', 'ratings']
    list_filter = ['teaching_cathedras']

    def full_name(self, obj):
        full_name = obj.user.get_full_name()
        return full_name if full_name else obj.user.username
    full_name.short_description = 'Юзер'

    def get_cathedras(self, obj):
        return ", ".join([p.cathedra for p in obj.teaching_cathedras.all()])
    get_cathedras.short_description = 'Кафедры'

    def get_report_button(self, obj):
        return format_html('<a class="button" href="{}">Сформировать отчет</a>', reverse('admin:user_grading_report', args=[obj.pk]))
    get_report_button.short_description = 'Отчет'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('report/<int:profile_id>/', self.admin_site.admin_view(self.user_grading_report), name='user_grading_report'),
        ]
        return custom_urls + urls

    def user_grading_report(self, request, profile_id):
        profile = self.get_object(request, profile_id)
        grading_records = Grading.objects.filter(user=profile)

        return render(request, 'admin/user_grading_report.html', {
            'profile': profile,
            'grading_records': grading_records,
        })

    def export_selected_profiles(self, request, queryset):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'User Reports'

        # Заполняем заголовки столбцов
        headers = ['ФИО', 'Наименование работ', 'Норматив в баллах', 'Выполненная работа', 'Баллы']
        sheet.append(headers)

        # Собираем данные и заполняем строки
        data = []
        for profile in queryset:
            # Фильтрация записей по статусу 'approved'
            grading_records = Grading.objects.filter(user=profile, status='approved')
            for grading in grading_records:
                row = [
                    profile.user.get_full_name(),
                    grading.used_standard.title,
                    grading.used_standard.standard_in_points,
                    grading.work_done,
                    grading.rating
                ]
                data.append(row)
                sheet.append(row)

        # Устанавливаем ширину колонок
        for col_num, col_name in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(
                len(str(cell)) for cell in [col_name] + [row[col_num-1] for row in data]
            )
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Создаем ответ с MIME-типом Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=selected_profiles_report.xlsx'

        # Сохраняем рабочий лист в ответе
        workbook.save(response)

        return response

    export_selected_profiles.short_description = "Экспорт выбранных профилей в Excel"

    actions = [export_selected_profiles]


class CriteriaAdmin(admin.ModelAdmin):
    list_display = ['title', 'standard_in_points', 'table_name']
    search_fields = ['title', 'standard_in_points', 'table_title__table']
    list_filter = ['table_title']

    def table_name(self, obj):
        return obj.table_title.table

    table_name.short_description = 'Название таблицы'


class GradingAdmin(admin.ModelAdmin):
    form = GradingForm
    list_display = ['full_name', 'criteria_title', 'work_done', 'rating', 'status']
    search_fields = ['user__user__first_name', 'user__user__last_name', 'used_standard__title', 'work_done', 'rating']
    list_filter = ['status']
    list_editable = ['status']

    def criteria_title(self, obj):
        return obj.used_standard.title

    def full_name(self, obj):
        # Достаем имя из связанного User объекта через Profile
        return obj.user.user.get_full_name()

    full_name.short_description = 'Юзер'
    criteria_title.short_description = 'Наименование работ'


class TableAdmin(admin.ModelAdmin):
    list_display = ['table']
    search_fields = ['table']


class FacultiesAdmin(admin.ModelAdmin):
    list_display = ['faculty']
    search_fields = ['faculty']


class CathedrasAdmin(admin.ModelAdmin):
    list_display = ['cathedra', 'owning_faculty']
    search_fields = ['cathedra', 'owning_faculty__faculty']
    list_filter = ['owning_faculty']


class InspectorsAdmin(admin.ModelAdmin):
    list_display = ['user', 'get_audited_faculties']
    search_fields = ['user__first_name', 'user__last_name', 'audited_faculty__faculty']
    list_filter = ['audited_faculty']

    def get_audited_faculties(self, obj):
        return ", ".join([faculty.faculty for faculty in obj.audited_faculty.all()])
    get_audited_faculties.short_description = 'Факультеты'


admin.site.register(Profile, ProfileAdmin)
admin.site.register(Criteria, CriteriaAdmin)
admin.site.register(Grading, GradingAdmin)
admin.site.register(Table, TableAdmin)
admin.site.register(Faculties, FacultiesAdmin)
admin.site.register(Cathedras, CathedrasAdmin)
admin.site.register(Inspectors, InspectorsAdmin)
